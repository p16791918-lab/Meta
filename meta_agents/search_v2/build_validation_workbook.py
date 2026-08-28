#!/usr/bin/env python3
"""Author validation workbook: makes the two human steps (full-text inclusion,
extraction verification) plus author review of RoB actually performable, and
adds a T/A-excluded random sample to estimate screening sensitivity.

Every "author" column is yellow; the AI's provisional values are locked reference.
A Summary sheet computes agreement rates by COUNTIF once the author fills them in.
"""
import csv, os, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "outputs", "Author_Validation_Workbook.xlsx")
def rd(p): return list(csv.DictReader(open(os.path.join(HERE, p), encoding="utf-8")))

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
YELLOW = PatternFill("solid", fgColor="FFF2CC")          # author-input columns
EX_FILL = PatternFill("solid", fgColor="E7E6E6")          # example row
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D9D9D9")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORD
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.row_dimensions[row].height = 34

def yn_validator(ws, col_letter, first, last, options='"Yes,No"'):
    dv = DataValidation(type="list", formula1=options, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("%s%d:%s%d" % (col_letter, first, col_letter, last))

def write_rows(ws, rows, widths, author_cols, example, wrapcols=()):
    """rows: list of lists (data). author_cols: 0-based indices of yellow input cols."""
    ncol = len(widths)
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
    # example row (row 2)
    for j, v in enumerate(example):
        cell = ws.cell(row=2, column=j + 1, value=v)
        cell.fill = EX_FILL; cell.font = Font(name=FONT, italic=True, size=9, color="7F7F7F")
        cell.alignment = WRAP if j in wrapcols else TOP; cell.border = BORD
    # data rows (row 3+)
    for r, row in enumerate(rows, start=3):
        for j, v in enumerate(row):
            cell = ws.cell(row=r, column=j + 1, value=v)
            cell.font = Font(name=FONT, size=9)
            cell.alignment = WRAP if j in wrapcols else TOP
            cell.border = BORD
            if j in author_cols:
                cell.fill = YELLOW
    ws.auto_filter.ref = "A1:%s%d" % (chr(64 + ncol), 2 + len(rows))
    return 2 + len(rows)  # last data row index

# ============================ Sheet 1: Full-text inclusion ============================
ws = wb.active; ws.title = "1_FullText_Inclusion"
inc = rd("TableS_included_studies.csv")
exc = rd("TableS_excluded_fulltext.csv")
headers = ["record_id", "Citation / title", "PMID", "DOI",
           "AI provisional decision", "AI reason",
           "AUTHOR decision (Include/Exclude)", "AUTHOR agrees with AI? (Yes/No)",
           "AUTHOR note / reason if changed"]
ws.append(headers); style_header(ws, len(headers))
rows = []
for r in inc:
    reason = "Included — %s synthesis; %s" % (r.get("synth_group", ""), r.get("data_source", ""))
    rows.append([r["record_id"], r.get("citation", ""), r.get("pmid", ""), r.get("doi", ""),
                 "Include", reason, "", "", ""])
for r in exc:
    det = (" — " + r["detail"]) if r.get("detail", "").strip() else ""
    rows.append([r["record_id"], r.get("citation", ""), r.get("pmid", ""), r.get("doi", ""),
                 "Exclude", r.get("exclusion_reason", "") + det, "", "", ""])
example = ["e.g. 1234", "Smith J, et al. Breast cancer incidence... Cancer. 2019.", "31234567",
           "10.1000/xyz", "Include", "Included — quantitative synthesis; SEER",
           "Include", "Yes", "(leave blank if you agree)"]
last1 = write_rows(ws, rows, [10, 46, 12, 22, 14, 34, 16, 16, 30],
                   author_cols=[6, 7, 8], example=example, wrapcols=(1, 5, 8))
yn_validator(ws, "G", 3, last1, '"Include,Exclude"')
yn_validator(ws, "H", 3, last1)
ws.cell(row=1, column=7).comment = Comment("Your own decision after reading the full text. "
    "This is the human eligibility judgment for the review.", "validation")

# ============================ Sheet 2: T/A excluded sample ============================
ws = wb.create_sheet("2_TA_Excluded_Sample")
dec = rd("screening_decisions.csv")
excluded = [r for r in dec if r["decision"] == "exclude"]
random.seed(20260828)
sample = random.sample(excluded, 150)
headers = ["record_id", "Title", "AI exclude reason",
           "AUTHOR: should this have been INCLUDED? (Yes/No)", "AUTHOR note"]
ws.append(headers); style_header(ws, len(headers))
rows = [[r["record_id"], r.get("title", ""), r.get("display_reason", ""), "", ""] for r in sample]
example = ["e.g. 5678", "Some title screened out at title/abstract stage",
           "Not relevant to the research question/topic", "No", "(note only if you disagree)"]
last2 = write_rows(ws, rows, [10, 60, 34, 20, 30], author_cols=[3, 4],
                   example=example, wrapcols=(1, 2, 4))
yn_validator(ws, "D", 3, last2)
ws.cell(row=1, column=4).comment = Comment("A random 150-record sample of the 4,551 records "
    "excluded at title/abstract. Re-reading these estimates how many eligible studies the "
    "AI first-pass screen may have missed (screening sensitivity).", "validation")

# ============================ Sheet 3: Extraction verification ============================
ws = wb.create_sheet("3_Extraction_Verification")
led = [r for r in rd("breast_extraction.csv") if r["record_id"] != "SEER-EXPL"]
headers = ["record_id", "Study (author-year)", "Group", "Comparator", "Outcome",
           "Minority rate [95% CI]", "NHW rate [95% CI]", "IRR [95% CI]",
           "Provenance", "Source location (table/figure)",
           "AUTHOR verified vs source? (Yes/No)", "AUTHOR corrected value (if wrong)", "AUTHOR note"]
ws.append(headers); style_header(ws, len(headers))
def ci(v, lo, hi):
    v = (v or "").strip()
    if not v: return ""
    return v + ((" [%s, %s]" % (lo, hi)) if (lo or "").strip() and (hi or "").strip() else "")
rows = []
for r in led:
    rows.append([r["record_id"],
                 r.get("author_year", ""), r.get("minority_group", ""), r.get("comparison_vs", ""),
                 r.get("outcome_dim", ""),
                 ci(r.get("minority_rate"), r.get("min_ci_lo"), r.get("min_ci_hi")),
                 ci(r.get("nhw_rate"), r.get("nhw_ci_lo"), r.get("nhw_ci_hi")),
                 ci(r.get("irr"), r.get("irr_ci_lo"), r.get("irr_ci_hi")),
                 r.get("provenance", ""), r.get("source_location", ""), "", "", ""])
example = ["e.g. 234", "Gomez2026_SEER21", "Chinese", "NHW", "disaggregated-AANHPI",
           "115.9 [113.5, 118.4]", "152.5 [151.9, 153.1]", "0.760 [0.745, 0.775]",
           "computed-from-rates-with-CI", "Table 2 / eTable 3", "Yes", "", "(matches source)"]
last3 = write_rows(ws, rows, [10, 20, 18, 14, 20, 22, 22, 20, 26, 24, 18, 20, 26],
                   author_cols=[10, 11, 12], example=example, wrapcols=(1, 8, 9, 12))
yn_validator(ws, "K", 3, last3)
ws.cell(row=1, column=10).comment = Comment("Where in the source the value came from — go here "
    "to confirm the extracted numbers.", "validation")

# ============================ Sheet 4: RoB review ============================
ws = wb.create_sheet("4_RoB_Review")
rob = rd("outputs/TableS_risk_of_bias.csv")
QC = [("Q1_frame", "Q1 frame"), ("Q2_sampling", "Q2 sampling"), ("Q3_size", "Q3 size"),
      ("Q4_described", "Q4 described"), ("Q5_coverage", "Q5 coverage"),
      ("Q6_condition", "Q6 condition"), ("Q7_measurement", "Q7 measure"),
      ("Q8_analysis", "Q8 analysis"), ("Q9_response", "Q9 response")]
headers = (["Study", "Registry", "Period"] + [lbl for _, lbl in QC] +
           ["AI Overall RoB", "AI justification",
            "AUTHOR agrees? (Yes/No)", "AUTHOR override RoB", "AUTHOR note"])
ws.append(headers); style_header(ws, len(headers))
rows = []
for r in rob:
    rows.append([r.get("study", ""), r.get("registry", ""), r.get("period", "")] +
                [r.get(k, "") for k, _ in QC] +
                [r.get("Overall_RoB", ""), r.get("justification", ""), "", "", ""])
example = (["Gomez2026_SEER21", "SEER-21", "2018-2022"] + ["Y"] * 9 +
           ["Low", "population-based registry; ...", "Yes", "", "(leave blank if you agree)"])
widths = [22, 14, 12] + [7] * 9 + [12, 40, 14, 14, 26]
acols = [len(headers) - 3, len(headers) - 2, len(headers) - 1]
last4 = write_rows(ws, rows, widths, author_cols=acols, example=example, wrapcols=(15,))
yn_validator(ws, chr(65 + len(headers) - 3), 3, last4)  # AUTHOR agrees column
yn_validator(ws, chr(65 + len(headers) - 2), 3, last4, '"Low,Moderate,High"')

# ============================ Sheet 5: Summary (agreement rates) ============================
ws = wb.create_sheet("5_Summary", 0)  # place first
ws.sheet_view.showGridLines = False
title = ws.cell(row=1, column=1, value="Author validation — agreement summary")
title.font = Font(name=FONT, bold=True, size=14, color="1F4E79")
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14
rowspec = [
    ("Full-text inclusion (Sheet 1)", None, None),
    ("  Studies to adjudicate", "=COUNTA('1_FullText_Inclusion'!A3:A400)", None),
    ("  Author decisions entered", "=COUNTA('1_FullText_Inclusion'!G3:G400)", None),
    ("  Author agrees with AI (Yes)", "=COUNTIF('1_FullText_Inclusion'!H3:H400,\"Yes\")", None),
    ("  Agreement rate", '=IFERROR(B6/B5,"")', "0.0%"),
    ("T/A excluded sample (Sheet 2)", None, None),
    ("  Records in sample", "=COUNTA('2_TA_Excluded_Sample'!A3:A400)", None),
    ("  Author reviewed", "=COUNTA('2_TA_Excluded_Sample'!D3:D400)", None),
    ("  Flagged 'should have been included' (Yes)", "=COUNTIF('2_TA_Excluded_Sample'!D3:D400,\"Yes\")", None),
    ("  Estimated false-exclusion rate", '=IFERROR(B11/B10,"")', "0.0%"),
    ("Extraction verification (Sheet 3)", None, None),
    ("  Estimates to verify", "=COUNTA('3_Extraction_Verification'!A3:A400)", None),
    ("  Author verified", "=COUNTIF('3_Extraction_Verification'!K3:K400,\"Yes\")+COUNTIF('3_Extraction_Verification'!K3:K400,\"No\")", None),
    ("  Confirmed correct (Yes)", "=COUNTIF('3_Extraction_Verification'!K3:K400,\"Yes\")", None),
    ("  Extraction agreement rate", '=IFERROR(B16/B15,"")', "0.0%"),
    ("Risk of bias review (Sheet 4)", None, None),
    ("  Studies to review", "=COUNTA('4_RoB_Review'!A3:A400)", None),
    ("  Author reviewed", "=COUNTIF('4_RoB_Review'!P3:P400,\"Yes\")+COUNTIF('4_RoB_Review'!P3:P400,\"No\")", None),
    ("  Author agrees (Yes)", "=COUNTIF('4_RoB_Review'!P3:P400,\"Yes\")", None),
    ("  RoB agreement rate", '=IFERROR(B21/B20,"")', "0.0%"),
]
r0 = 3
for i, (label, formula, numfmt) in enumerate(rowspec):
    rr = r0 + i
    a = ws.cell(row=rr, column=1, value=label)
    if formula is None:  # section header
        a.font = Font(name=FONT, bold=True, size=11, color="1F4E79")
    else:
        a.font = Font(name=FONT, size=10)
        b = ws.cell(row=rr, column=2, value=formula)
        b.font = Font(name=FONT, size=10, bold=("rate" in label))
        if numfmt: b.number_format = numfmt
ws.cell(row=r0 + len(rowspec) + 1, column=1,
        value="Rates populate automatically as you fill the yellow columns in each sheet.").font = \
    Font(name=FONT, italic=True, size=9, color="7F7F7F")

# ============================ README ============================
ws = wb.create_sheet("README", 0)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 110
lines = [
    ("Author validation workbook — breast cancer SR", 15, True, "1F4E79"),
    ("", 10, False, None),
    ("Purpose: make the human-performed steps of the review real and documented. Fill the YELLOW columns only.", 10, False, None),
    ("The AI performed title/abstract screening and data extraction; the author performs full-text inclusion", 10, False, None),
    ("decisions and verifies the extracted values and risk-of-bias ratings. Complete every sheet, then the", 10, False, None),
    ("Summary sheet reports the agreement rates to cite in Methods.", 10, False, None),
    ("", 10, False, None),
    ("Sheet 1 — Full-text inclusion (242 reports): read each full text and record YOUR Include/Exclude", 10, True, None),
    ("        decision (col G) and whether you agree with the AI (col H). This is the human eligibility step.", 10, False, None),
    ("Sheet 2 — T/A excluded sample (150 records): re-read this random sample of records the AI excluded at", 10, True, None),
    ("        title/abstract; flag any that should have been included (col D). Estimates screening sensitivity.", 10, False, None),
    ("Sheet 3 — Extraction verification (147 estimates): open each source at the listed location and confirm", 10, True, None),
    ("        the extracted numbers (col K); enter a correction if any value is wrong (col L).", 10, False, None),
    ("Sheet 4 — Risk of bias review (43 studies): confirm or override each JBI rating (cols P–Q).", 10, True, None),
    ("Sheet 5 — Summary: agreement rates, computed automatically. Nothing to fill here.", 10, True, None),
    ("", 10, False, None),
    ("Row 2 of each sheet is a grey EXAMPLE row showing the expected format — delete it before finalizing.", 10, False, "C00000"),
    ("Yellow = you fill.  White = AI reference (do not edit).  Dropdowns are provided where applicable.", 10, False, None),
]
for i, (txt, sz, bold, color) in enumerate(lines, start=1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(name=FONT, size=sz, bold=bold, color=color or "000000")

# order: README, Summary, then the four working sheets
wb.move_sheet("README", -wb.sheetnames.index("README"))
wb.save(OUT)
print("wrote", OUT)
print("counts: fulltext=%d (inc %d/exc %d), TA-sample=150, extraction=%d, RoB=%d"
      % (len(inc) + len(exc), len(inc), len(exc), len(led), len(rob)))
